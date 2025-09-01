import argparse
import configparser
import datetime
import requests
from openpyxl import Workbook

def get_config():
    """Reads the configuration from config.ini."""
    config = configparser.ConfigParser()
    config.read('config.ini')
    return config

def get_client_ids(config):
    """Reads the client IDs from the config."""
    client_ids_str = config.get('clients', 'ids', fallback='')
    if not client_ids_str:
        return []
    return [int(id.strip()) for id in client_ids_str.split(',')]

def get_previous_month_dates():
    """Gets the start and end dates for the previous month."""
    today = datetime.date.today()
    first_day_of_current_month = today.replace(day=1)
    last_day_of_previous_month = first_day_of_current_month - datetime.timedelta(days=1)
    first_day_of_previous_month = last_day_of_previous_month.replace(day=1)
    return first_day_of_previous_month, last_day_of_previous_month

def get_client_name(config, client_id):
    """Fetches the client name from the Harvest API."""
    token = config.get('harvest', 'token')
    account_id = config.get('harvest', 'account_id')
    headers = {
        "Authorization": f"Bearer {token}",
        "Harvest-Account-ID": account_id,
        "User-Agent": "Python Harvest Reporter"
    }
    url = f"https://api.harvestapp.com/v2/clients/{client_id}"
    response = requests.get(url, headers=headers)
    response.raise_for_status()
    return response.json()['name']

def get_time_entries(config, client_id, from_date, to_date):
    """Fetches time entries for a client from the Harvest API."""
    token = config.get('harvest', 'token')
    account_id = config.get('harvest', 'account_id')
    headers = {
        "Authorization": f"Bearer {token}",
        "Harvest-Account-ID": account_id,
        "User-Agent": "Python Harvest Reporter"
    }
    url = f"https://api.harvestapp.com/v2/time_entries"
    params = {
        "client_id": client_id,
        "from": from_date.isoformat(),
        "to": to_date.isoformat()
    }
    response = requests.get(url, headers=headers, params=params)
    response.raise_for_status()  # Raise an exception for bad status codes
    return response.json()['time_entries']

def create_report(client_name, time_entries, from_date, to_date=None):
    """Creates an Excel report for a client."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Remote"

    # Write header
    ws.append(["Datum", "Leistung", "Arbeitsstunden"])

    row_num = 2
    for entry in time_entries:
        ws.append([
            entry['spent_date'],
            entry['notes'],
            entry['hours']
        ])
        row_num += 1

    # Write summary
    ws.append([])
    summary_row = row_num + 1
    ws.cell(row=summary_row, column=1, value="SUMME")
    ws.cell(row=summary_row, column=2, value=f"=SUM(C2:C{row_num-1})")

    # Save the workbook
    if to_date:
        date_str = f"{from_date.strftime('%Y-%m-%d')}_to_{to_date.strftime('%Y-%m-%d')}"
    else:
        date_str = from_date.strftime('%B_%Y')
    report_filename = f"Leistungsuebersicht {client_name} {date_str}.xlsx"
    wb.save(report_filename)
    print(f"Report saved to {report_filename}")

def main():
    """Main function to generate reports."""
    parser = argparse.ArgumentParser(description="Generate Harvest time reports.")
    parser.add_argument("--from_date", help="Start date for the report in YYYY-MM-DD format.")
    parser.add_argument("--to_date", help="End date for the report in YYYY-MM-DD format.")
    args = parser.parse_args()

    config = get_config()
    client_ids = get_client_ids(config)

    if not client_ids:
        print("No client IDs found in config.ini. Exiting.")
        return

    if args.from_date and args.to_date:
        from_date = datetime.datetime.strptime(args.from_date, "%Y-%m-%d").date()
        to_date = datetime.datetime.strptime(args.to_date, "%Y-%m-%d").date()
        print(f"Generating reports from {from_date} to {to_date}...")
    else:
        from_date, to_date = get_previous_month_dates()
        print(f"Generating reports for {from_date.strftime('%B %Y')}...")

    for client_id in client_ids:
        try:
            client_name = get_client_name(config, client_id)
            print(f"Fetching time entries for client {client_name} ({client_id})...")
            time_entries = get_time_entries(config, client_id, from_date, to_date)
            if time_entries:
                if args.from_date and args.to_date:
                    create_report(client_name, time_entries, from_date, to_date)
                else:
                    create_report(client_name, time_entries, from_date)
            else:
                print(f"No time entries found for client {client_name} for the specified period.")
        except requests.exceptions.RequestException as e:
            print(f"Error fetching data for client {client_id}: {e}")
        except Exception as e:
            print(f"An error occurred while processing client {client_id}: {e}")

if __name__ == "__main__":
    main()
